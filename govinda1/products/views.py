from django.shortcuts import render
from django.http import HttpResponse
from django.contrib.auth import authenticate, login,	logout
from django.contrib.auth.decorators import login_required
from	.models	import *
from django.utils import timezone
from	django.shortcuts	import	render,	get_object_or_404
from django.db.models import Q
from django.db.models import Sum
from datetime import datetime, timedelta
from django.core.management import call_command
import io
from django.http import FileResponse
import json
#from reportlab.pdfgen import canvas
#from reportlab.lib.units import inch
#from reportlab.lib.pagesizes import letter, A5, A4
#from reportlab.platypus.tables import Table,TableStyle,colors
#from reportlab.platypus import SimpleDocTemplate
#from reportlab.lib.pagesizes import letter
#from reportlab.platypus import Table
from django.conf import settings
from django.template.loader import get_template
from xhtml2pdf import pisa
from django.contrib.staticfiles import finders


import	datetime
import	csv
import	os
import	subprocess
import	openpyxl
import	functools
import	pdb
import logging
import	re
import requests
import	json
import time
import random
import string

# Create your views here.

def index(request):
	#return HttpResponse("Hello Worls ")
	return render(request,'products/MenuList.html')

@login_required
def list_main(request):
	return render(request,'products/MenuList.html')

def	login_user(request):
							error=""
							if	request.POST:
										uid=request.POST['uid']
										pwd	=request.POST['pwd']
										user1 = authenticate(username=uid, password=pwd)
										if	user1	is	not	None:
																login(request,user1)
																
																return render(request,"products/MenuList.html", {})
										else:
														error="Invalid	Credentials"
							return	render(request,'products/Login.html',{'error':error})

def	logout_user(request):
				logout(request)
				success="Successfully"
				return	render(request,'products/Login.html',{'success':success})

@login_required
def list_alerts(request):
	
		return render(request,'products/ListAlerts.html',{})

@login_required
def list_groups(request):

		return render(request,'products/ListGroups.html',{'inventorygroup':InventoryGroups.objects.all()})


@login_required
def	export_host_list(request):
					response=HttpResponse(content_type='text/csv')
					response['Content-Disposition']	=	'attachment;	filename=PowerVCHosts'	+	str(datetime.datetime.now())	+	'.csv'
					writer=csv.writer(response)
					writer.writerow(['FSP_/_BMC_IP','Serial','Location','CEC_Name','FSP_Credentials',	'Host_Type',	'Firmware_Level',	'HMC_IP',	'Novalink_IP',	'VIOS1',	'VIOS2',	'Proc_/_Memory',	'Model',	'Network_Ports',	'Fabric_Switch1',	'Fabric_Switch2',	'Primary_VLAN',	'Connected_Storages',	'Owner',	'Squad',	'LAB'])
					hosts={}
					for	h	in	hosts:
								writer.writerow([])
					return	response


@login_required
def list_bu(request):
		return render(request,'products/ListBU.html',{'bus':BusinessUnit.objects.all()})



@login_required
def add_bu(request):
		success=""
		error=""
		ig=""
		now1=timezone.now()
		if	request.POST:
								if	'bu_name'	in	request.POST:
										if	str(request.POST['bu_name']).strip():
												try:
													ig= BusinessUnit.objects.get(pk=str(request.POST['bu_name']).strip())
													error="Business Unit with the same name already present"
												except	BusinessUnit.DoesNotExist:	
															bname=str(request.POST['bu_name']).strip()
															bdesc=request.POST['bu_desc']
															bdate= now1
															bu=BusinessUnit(name=bname,desc=bdesc,date=bdate)
															bu.save()
															success="BU	is	successfully	created"
										else:
															error="Mandatory	fields	are	missing"

		return render(request,'products/AddBU.html',{'now':now1,	'psuccess':success,	'perror':error})

@login_required
def edit_bu(request):
						perror=""
						psuccess=""
						eb=""
						bu=""
						if	request.POST:
									if	'edit_bu'	in	str(request.POST):
												try:
																eb = BusinessUnit.objects.get(pk=str(request.POST['edit_bu']).strip())
																return render(request,'products/EditBU.html',{'bu':eb})
												except	BusinessUnit.DoesNotExist:																																	
																perror="Given BU Doesn't exist"

									elif	'bu_name'	in	str(request.POST):
											if 'bu_desc' in	str(request.POST) :			
												bu=get_object_or_404(BusinessUnit,	pk=str(request.POST['bu_name']).strip())
												bu.desc=request.POST['bu_desc']
												bu.save()
												psuccess="Successfully Edited the Business Unit"
											
						return render(request,'products/EditBU.html',{'perror':perror,'psuccess':psuccess})

@login_required
def delete_bu(request):
						perror=""
						psuccess=""
						db=""
						bu=""
						if	request.POST:
									if	'del_bu'	in	str(request.POST):
												try:
																db = BusinessUnit.objects.get(pk=str(request.POST['del_bu']).strip())
																return render(request,'products/DeleteBU.html',{'dbu':db})
												except	BusinessUnit.DoesNotExist:																																	
																perror="Given BU Doesn't exist"

									elif	'bu_name'	in	str(request.POST):
												db=get_object_or_404(BusinessUnit,	pk=str(request.POST['bu_name']).strip())
												db.delete()
												psuccess="Successfully Deleted the Business Unit"
											
						return render(request,'products/DeleteBU.html',{'perror':perror,'psuccess':psuccess})



@login_required
def add_group(request):
				success=""
				error=""
				ig=""
				now1=timezone.now()
				bus=BusinessUnit.objects.all()
				if	request.POST:
								if	'grp_name'	in	request.POST:
										if	str(request.POST['grp_name']).strip() and str(request.POST['bu']).strip() !='----':
												try:
													ig= InventoryGroups.objects.get(pk=str(request.POST['grp_name']).strip())
													error="Inventory Group with the same name already present"
												except	InventoryGroups.DoesNotExist:	
															gname=str(request.POST['grp_name']).strip()
															gdesc=request.POST['grp_desc']
															gdate= now1
															bu=request.POST['bu']
															ig=InventoryGroups(name=gname,desc=gdesc,date=gdate,bu=BusinessUnit.objects.get(pk=bu))
															ig.save()
															success="Group	is	successfully	created"
										else:
															error="Mandatory	fields	are	missing"
				return render(request,'products/AddGroup.html',{'now':now1,'bus':bus, 'psuccess':success, 'perror':error})


@login_required
def edit_group(request):
						perror=""
						psuccess=""
						eg=""
						ig=""				
						bus=BusinessUnit.objects.all()
						if	request.POST:
									if	'edit_gr'	in	str(request.POST):
												try:
																eg = InventoryGroups.objects.get(pk=str(request.POST['edit_gr']).strip())
																return render(request,'products/EditGroup.html',{'bus':bus,'ig':eg})
												except	InventoryGroups.DoesNotExist:																																	
																perror="Given Inventory Group Doesn't exist"

									elif	'gr_name'	in	str(request.POST):
											if 'gr_desc' in	str(request.POST) :			
												ig=get_object_or_404(InventoryGroups,	pk=str(request.POST['gr_name']).strip())
												ig.desc=request.POST['gr_desc']
												ig.bu=BusinessUnit.objects.get(pk=str(request.POST['bu']).strip())
												ig.save()
												psuccess="Successfully Edited the Inventory group"
											
						return render(request,'products/EditGroup.html',{'bus':bus,'perror':perror,'psuccess':psuccess})



@login_required
def delete_group(request):
						perror=""
						psuccess=""
						dg=""
						ig=""
						if	request.POST:
									if	'delete_gr'	in	str(request.POST):
												try:
																dg = InventoryGroups.objects.get(pk=str(request.POST['delete_gr']).strip())
																return render(request,'products/DeleteGroup.html',{'dg':dg,'perror':perror,'psuccess':psuccess})
												except	InventoryGroups.DoesNotExist:																																	
																perror="Given Inventory Group Doesn't exist"

									elif 'gr_name'	in	str(request.POST):
												ig=get_object_or_404(InventoryGroups,	pk=str(request.POST['gr_name']).strip())
												ig.delete()
												psuccess="Successfully Deleted the Inventory group"
											
						return render(request,'products/DeleteGroup.html',{'perror':perror,'psuccess':psuccess})

@login_required
def list_inventory(request):
	qs=""
	qs=InventoryItems.objects.all()
	bus=BusinessUnit.objects.all()
	bu_sel=request.GET.get('bu')
	item_g=request.GET.get('item_group')
	item_n=request.GET.get('item_name')
	item_c=request.GET.get('item_code')

	if bu_sel !="-----" and bu_sel is not None:
			qs=qs.filter(bu=get_object_or_404(BusinessUnit,	pk=str(bu_sel).strip()))
	if item_g !="-----" and item_g is not None:
			qs=qs.filter(group=get_object_or_404(InventoryGroups,	pk=str(item_g).strip()))
	if item_n !="" and item_n is not None:
			qs=qs.filter(name__icontains=item_n)
	if item_c !="" and item_c is not None:
			qs=qs.filter(item_code__icontains=item_c)
	
	return render(request,'products/ListInventoryItem.html',{'bus':bus,'qs':qs})

@login_required
def edit_cart(request):
		perror=""
		psuccess=""
		cart_list=""
		order_list=""
		edit="Yes"
		bus=BusinessUnit.objects.all()

		if Cart.objects.count()==0:
			edit=""
		if request.POST:
				edit=""
				if 'i_code' in request.POST and request.POST.get('i_code') != "":
							if 'order' in request.POST:
								cart_list=request.POST.getlist('i_code')
								order_list=request.POST.getlist('order')
								i=0
								while i < len(order_list):
									o_item=""
									c_item=""
									c_item=Cart.objects.get(pk=str(cart_list[i]).strip())
									o_item=InventoryItems.objects.get(pk=str(cart_list[i]).strip())
									
									if order_list[i] != c_item.order_quantity:
										o_item.available_quantity+=(c_item.order_quantity-int(order_list[i]))
										o_item.save()
										if int(order_list[i])==0:
											c_item.delete()
										else:
											c_item.order_quantity=int(order_list[i])
											c_item.order_price=round(c_item.order_quantity * c_item.item_net_price,2)
											c_item.save()
										
									
									i+=1
		gst=0
		iprice=0
		total=0
		if Cart.objects.exists():
			for j in Cart.objects.all():
				gst=gst+(j.tax*j.order_quantity)
				iprice=iprice+(j.item_price*j.order_quantity)
			total=round(Cart.objects.aggregate(Sum('order_price'))['order_price__sum'],2)
		return render(request,'products/AddSalesItem.html',{'bus':bus,'qs':InventoryItems.objects.all(),'edit':edit,'iprice':round(iprice,2),'tax':round(gst,2),'total':total,'cart':Cart.objects.all().order_by('bu')})


@login_required
def scan_items(request):
	# import cv2
	# from pyzbar.pyzbar import decode
	# from pybud import AudioSegment
	# from pydub.playback import play



	# cap = cv2.videocapture(0)

	# song= AudioSegment.from_wav("beep-02.wav")
	# while cap.isopened():
	# 	success, frame = cap.read()
	# 	frame=cv2.flip(frame,1)
	# 	detectedBarcode= decode(frame)
	# 	if not detectedBarcode:
	# 		print("No Barcode Detected")
	# 	else:
	# 		for barcode in detectedBarcode:
	# 			if barcode.data !="":
	# 				cv2.putText(frame,str(barcode.data),(50,50), cv2.FONT_HERSHEY_COMPLEX,2,(0,255,255),2)
	# 				play(song)
	# 				cv2.imwrite("code.png",frame)
	# 				print(barcode.data)
					
	# 	cv2.imshow("scanner",frame)
	# 	if cv2.waitkey(1)== ord('q'):
	# 		break
		# while true:
		# 	text = input()
		# 	print (text)


		######################<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<< NEW CODE TO CHECK
	# value1=""
	# #import pdb; pdb.set_trace();
	# while True:
	# 	if request.POST:
	# 		update_cart(request,value1)
	# 	else:
	# 		value1=input()
	# 	if value1 in ["q","exit"]:
	# 		break
	# 	else:
	# 		print ("I'm in Scan function "+ value1 )
	# 		update_cart(request,value1)
		import cv2
		from pyzbar.pyzbar import decode
		from pybud import AudioSegment
		from pydub.playback import play



		cap = cv2.videocapture(0)

		song= AudioSegment.from_wav("beep-02.wav")
		while cap.isopened():
			success, frame = cap.read()
			frame=cv2.flip(frame,1)
			detectedBarcode= decode(frame)
			if not detectedBarcode:
				print("No Barcode Detected")
			else:
				for barcode in detectedBarcode:
					if barcode.data !="":
						cv2.putText(frame,str(barcode.data),(50,50), cv2.FONT_HERSHEY_COMPLEX,2,(0,255,255),2)
						play(song)
						cv2.imwrite("code.png",frame)
						print(barcode.data)
						
			cv2.imshow('scanner',frame)
			if cv2.waitkey(1)== ord('q'):
				break



def update_cart(request):
			perror=""
			psuccess=""
			qs=""
			gst=0
			bus=""
			bu_sel=""
   
			item_g=""
			item_n=""
			item_c=""
			
			#print("I'm in Update cart, got the item id "+ value1)
			#import pdb; pdb.set_trace();
			bus=BusinessUnit.objects.all()
			qs=InventoryItems.objects.all()
			if request.GET:
					bu_sel=request.GET.get('bu')
					item_g=request.GET.get('item_group')
					item_n=request.GET.get('item_name')
					item_c=request.GET.get('item_code')

					if bu_sel !="-----" and bu_sel is not None:
							qs=qs.filter(bu=get_object_or_404(BusinessUnit,	pk=str(bu_sel).strip()))
					if item_g !="-----" and item_g is not None:
							qs=qs.filter(group=get_object_or_404(InventoryGroups,	pk=str(item_g).strip()))
					if item_n !="" and item_n is not None:
							qs=qs.filter(name__icontains=item_n)
					if item_c !="" and item_c is not None:
							qs=qs.filter(item_code__icontains=item_c)
					#import pdb; pdb.set_trace();
					#print("GOVINDA HARI >>>> "+str(request.POST.get('i_code')))
					
					#print(str(request.POST.get('iu_code')))

					# import pdb; pdb.set_trace();
			elif request.POST:
					#import pdb; pdb.set_trace();
					cart_list=[]
					order_list=[]
					a=0
					#import pdb; pdb.set_trace();
					inv_list=""
					inv_list=InventoryItems.objects.values_list('item_code',flat=True)
					if  ('barcode-input' in request.POST and request.POST.get('barcode-input') != ""):
							#import pdb; pdb.set_trace();
							bcode=""
							bcode=request.POST.get('barcode-input')
							if (bcode in inv_list):					
								c_item=""
								o_item=InventoryItems.objects.get(pk=str(bcode).strip())
								c_list=[]
								c_list=Cart.objects.values_list('item_code',flat=True)
								if (o_item.available_quantity >0):
									if  (bcode in c_list):
										c_item=Cart.objects.get(pk=str(bcode).strip())
									#except	Cart.DoesNotExist:
									else:
										c_item=Cart(item_code=o_item.item_code,name=o_item.name,unit=o_item.unit,order_quantity=0,item_price=o_item.price,igst=o_item.igst,cgst=o_item.cgst,sgst=o_item.sgst,tax=0,exp_date=o_item.exp_date,damage=o_item.damage,item_net_price=0,order_price=0,bu=str(o_item.bu.name),item_group=str(o_item.group.name))
									#import pdb; pdb.set_trace();
									c_item.order_quantity=c_item.order_quantity+1
									if c_item.igst > 0 or c_item.cgst > 0 or c_item.sgst > 0:
										c_item.tax=round(((c_item.igst+c_item.cgst+c_item.sgst)*c_item.item_price)/100,2)
									else:
										c_item.tax=0
									c_item.item_net_price=round(c_item.item_price+c_item.tax,2)
									c_item.order_price=round(c_item.order_quantity * c_item.item_net_price,2)
									c_item.save()
			
									
									o_item.available_quantity=o_item.available_quantity-1
									o_item.save()

					elif ('iu_code' in request.POST and request.POST.get('iu_code') != ""):
							
							cart_list=request.POST.getlist('iu_code')
							if 'order' in request.POST:
								order_list=request.POST.getlist('order')
							if len(cart_list) == len(order_list):
								c_items={}
								j=0
								for item in cart_list:
									if int(order_list[j])>0:
										c_items[item]=order_list[j]
									j=j+1
								x=""
								y=""
								for x,y in c_items.items():
										o_item=InventoryItems.objects.get(pk=str(x).strip())
										try:
											c_item=Cart.objects.get(pk=str(x).strip())
										except	Cart.DoesNotExist:
												c_item=Cart(item_code=o_item.item_code,name=o_item.name,unit=o_item.unit,order_quantity=0,item_price=o_item.price,igst=o_item.igst,cgst=o_item.cgst,sgst=o_item.sgst,tax=0,exp_date=o_item.exp_date,damage=o_item.damage,item_net_price=0,order_price=0,bu=str(o_item.bu.name),item_group=str(o_item.group.name))
										c_item.order_quantity+=int(y)
										if c_item.igst > 0 or c_item.cgst > 0 or c_item.sgst > 0:
											c_item.tax=round(((c_item.igst+c_item.cgst+c_item.sgst)*c_item.item_price)/100,2)
										else:
											c_item.tax=0
										c_item.item_net_price=round(c_item.item_price+c_item.tax,2)
										c_item.order_price=round(c_item.order_quantity * c_item.item_net_price,2)
										c_item.save()
										o_item.available_quantity-=int(y)
										o_item.save()
							#import pdb; pdb.set_trace();
					# flag=0
					# inv_list=""
					# inv_list=InventoryItems.objects.values_list('item_code',flat=True)
					# if (len(inv_list)>0) :
					# 	if ((len(cart_list)==1) and (cart_list[0] in inv_list)) or len(cart_list)>1:
					# 		flag=1

					# if flag==1:
					# 	i=0
					# 	while i < len(order_list):
					# 		o_item=""
					# 		c_item=""
					# 		if(int(order_list[i])>0):
					# 			o_item=InventoryItems.objects.get(pk=str(cart_list[i]).strip())
					# 			o_item.available_quantity-=int(order_list[i])
					# 			o_item.save()
					# 			try:
					# 				c_item=Cart.objects.get(pk=str(cart_list[i]).strip())
					# 			except	Cart.DoesNotExist:
					# 					import pdb; pdb.set_trace();
					# 					c_item=Cart(item_code=o_item.item_code,name=o_item.name,unit=o_item.unit,order_quantity=0,item_price=o_item.price,igst=o_item.igst,cgst=o_item.cgst,sgst=o_item.sgst,tax=0,exp_date=o_item.exp_date,damage=o_item.damage,item_net_price=0,order_price=0,bu=str(o_item.bu.name),item_group=str(o_item.group.name))
					# 			import pdb; pdb.set_trace();
					# 			c_item.order_quantity+=int(order_list[i])
					# 			if c_item.igst > 0 or c_item.cgst > 0 or c_item.sgst > 0:
					# 				c_item.tax=round(((c_item.igst+c_item.cgst+c_item.sgst)*c_item.item_price)/100,2)
					# 			else:
					# 				c_item.tax=0
					# 			c_item.item_net_price=round(c_item.item_price+c_item.tax,2)
					# 			c_item.order_price=round(c_item.order_quantity * c_item.item_net_price,2)
					# 			c_item.save()

					# 		i=i+1
			gst=0
			iprice=0
			total=0
			cart_list=""
			if Cart.objects.exists():
				for j in Cart.objects.all():
					gst=gst+(j.tax*j.order_quantity)
					iprice=iprice+(j.item_price*j.order_quantity)
				total=round(Cart.objects.aggregate(Sum('order_price'))['order_price__sum'],2)
				cart_list=Cart.objects.all().order_by('bu')
			return render(request,'products/AddSalesItem.html',{'qs':qs,'bus':bus,'total':total,'iprice':str(round(iprice,2)),'tax':str(round(gst,2)),'cart':cart_list,'perror':perror,'psuccess':psuccess})
			
@login_required
def clear_cart(request):
			perror=""
			psuccess=""
			item=""
			qs=Cart.objects.all()
			for cart_item in qs:
					try:
						item=InventoryItems.objects.get(pk=str(cart_item.item_code).strip()) 
						item.available_quantity=item.available_quantity+cart_item.order_quantity
						item.save()
						cart_item.delete()
					except:
						perror+="Cart Item not found in the inventory "+cart_item.item_code
      
			Cart.objects.all().delete()
			return render(request,'products/AddSalesItem.html',{'cart':Cart.objects.all().order_by('bu'),'qs':InventoryItems.objects.all(),'ig':InventoryGroups.objects.all(),'perror':perror,'psuccess':psuccess})

@login_required
def add_inventory(request):
				success=""
				error=""
				item=""
				now1=timezone.now()
				bus=BusinessUnit.objects.all()
				#import pdb; pdb.set_trace();
				if	request.POST:
								if	'item_code'	in	request.POST:
										if	(str(request.POST['item_code']).strip()) and (str(request.POST['bu']).strip()!="-----") and (str(request.POST['item_group']).strip()!=""):
												try:
													item= InventoryItems.objects.get(pk=str(request.POST['item_code']).strip())
													error="Inventory Item with the same Item code  already present"
												except	InventoryItems.DoesNotExist:
															#import pdb; pdb.set_trace();
															item_code=str(request.POST['item_code']).strip()
															name=str(request.POST['item_name']).strip()
															unit=str(request.POST['unit']).strip()
															if (request.POST['quantity']).strip():
																item_quantity= int(str(request.POST['quantity']).strip())
															else:
																item_quantity=0
															if (request.POST['price']).strip():
																	item_price=float(str(request.POST['price']).strip())
															else:
																item_price=0

															if (request.POST['igst']).strip():
																igst=float(str(request.POST['igst']).strip())
															else:
																igst=0
															if (request.POST['cgst']).strip():
																cgst=float(str(request.POST['cgst']).strip())
															else:
																cgst=0
															if (request.POST['sgst']).strip():
																sgst=float(str(request.POST['sgst']).strip())
															else:
																sgst=0
															date=now1
															exp_date=""
															#import pdb; pdb.set_trace();
															if  (request.POST['exp_date']).strip():
																exp_date=(str(request.POST['exp_date']).strip())
															else:
																exp_date="1111-11-11"
															damage_status=False
															if (request.POST['damage']).strip()=='Y':
																damage_status=True
															bu=get_object_or_404(BusinessUnit,	pk=str(request.POST['bu']).strip())
															group=get_object_or_404(InventoryGroups,	pk=str(request.POST['item_group']).strip())
															item=InventoryItems(item_code=item_code, name=name, unit=unit, available_quantity=item_quantity,price=item_price,igst=igst,cgst=cgst,sgst=sgst,date=date,exp_date=exp_date,bu=bu,group=group,damage=damage_status)
															item.save()
															success="Inventory Item	is	successfully Added"
										else:
															error="Mandatory	fields	are	missing"
				return render(request,'products/AddInventoryItem.html',{'now':now1,'bus':bus,'ig':InventoryGroups.objects.values_list('name',flat=True),	'psuccess':success,	'perror':error})


@login_required
def edit_inventory(request):
						perror=""
						psuccess=""
						item=""
						bus=BusinessUnit.objects.all()
						if	request.POST:
									if	'edit_item'	in	str(request.POST):
												try:
																item = InventoryItems.objects.get(pk=str(request.POST['edit_item']).strip())
																return render(request,'products/EditInventoryItem.html',{'item':item,'bus':bus})
												except	InventoryItems.DoesNotExist:																																	
																perror="Given Inventory Item Doesn't exist"

									elif	'item_code'	in	str(request.POST) and (str(request.POST['bu']).strip()!="-----") and (str(request.POST['item_group']).strip()!=""):
												

												item=get_object_or_404(InventoryItems,	pk=str(request.POST['item_code']).strip())
												item.name=str(request.POST['item_name']).strip()
												item.unit=str(request.POST['unit']).strip()
												item.available_quantity= int(str(request.POST['quantity']).strip())
												item.price=float(str(request.POST['price']).strip())
												item.igst=float(str(request.POST['igst']).strip())
												item.cgst=float(str(request.POST['cgst']).strip())
												item.sgst=float(str(request.POST['sgst']).strip())
												if str(request.POST['exp_date']).strip():
													item.exp_date=(str(request.POST['exp_date']).strip())
												else:
													item.exp="1111-11-11"
												item.bu=get_object_or_404(BusinessUnit,	pk=str(request.POST['bu']).strip())
												item.group=get_object_or_404(InventoryGroups,	pk=str(request.POST['item_group']).strip())
												damage_status=False
												if (request.POST['damage']).strip()=='Y':
													damage_status=True
												item.damage=damage_status
												#import pdb; pdb.set_trace()

												item.save()
												psuccess="Inventory Item	is	successfully Edited"
												
									else:
											perror="Mandatory fields are missing"	

						return render(request,'products/EditInventoryItem.html',{'bus':bus,'perror':perror,'psuccess':psuccess})

@login_required
def delete_inventory(request):
						perror=""
						psuccess=""
						item=""
						if	request.POST:
									if	'delete_item'	in	str(request.POST):
												try:
																item = InventoryItems.objects.get(pk=str(request.POST['delete_item']).strip())
																return render(request,'products/DeleteInventoryItem.html',{'ditem':item,'perror':perror,'psuccess':psuccess})
												except	InventoryItems.DoesNotExist:																																	
																perror="Given Inventory Item Doesn't exist"

									elif	'item_code'	in	str(request.POST):
												item=get_object_or_404(InventoryItems,	pk=str(request.POST['item_code']).strip())
												item.delete()
												psuccess="Inventory Item is	successfully Deleted"
												
											
						return render(request,'products/DeleteInventoryItem.html',{'perror':perror,'psuccess':psuccess})

@login_required
def load_groups(request):
		#import pdb; pdb.set_trace()
		ig=""
		bu=request.GET.get('bu')
		if bu =="-----":
			igl=[]
		else:
			igl=InventoryGroups.objects.filter(bu=BusinessUnit.objects.get(pk=bu)).values_list('name',flat=True)
		return render(request, 'products/groups_list.html', {'igl':igl})

@login_required
def list_customers(request):
	perror=""
	psuccess=""
	qs=[]
	mobile=request.GET.get('mobile')
	name=request.GET.get('name')
	mail=request.GET.get('mail')
	addr=request.GET.get('address')
	if mobile or name or mail or addr :
		qs=Customers.objects.all()
		if mobile !="" and mobile is not None:
				qs=qs.filter(customer_mobile__icontains=mobile)
		if name !="" and name is not None:
				qs=qs.filter(customer_name__icontains=name)
		if mail !="" and mail is not None:
				qs=qs.filter(customer_mailid__icontains=mail)
		if addr !="" and addr is not None:
				qs=qs.filter(customer_address__icontains=addr)


	return render(request,'products/ListCustomers.html',{'customers':qs,'perror':perror,'psuccess':psuccess})

@login_required
def edit_customer(request):
						perror=""
						psuccess=""
						ec=""
						bu=""
						if	request.POST:
									if	'edit_cust'	in	str(request.POST):
												try:
																ec = Customers.objects.get(pk=str(request.POST['edit_cust']).strip())
																return render(request,'products/EditCustomer.html',{'ecust':ec})
												except	Customers.DoesNotExist:																																	
																perror="Given Customer Mobile number Doesn't exist"
									elif  'ecust'	in	str(request.POST):
												c_name=str(request.POST['ecust_name']).strip()
												c_mail=str(request.POST['ecust_mail']).strip()
												c_addr=str(request.POST['ecust_addr']).strip()

												ec=get_object_or_404(Customers,	pk=str(request.POST['ecust']).strip())
												ec.customer_name=c_name
												ec.customer_mailid=c_mail
												ec.customer_address=c_addr

												ec.save()
												psuccess="Successfully Edited the Customer info"
											
						return render(request,'products/EditCustomer.html',{'perror':perror,'psuccess':psuccess})
		

@login_required
def delete_customer(request):
						perror=""
						psuccess=""
						dc=""
						bu=""
						if	request.POST:
									if	'del_cust'	in	str(request.POST):
												try:
																dc = Customers.objects.get(pk=str(request.POST['del_cust']).strip())
																return render(request,'products/DeleteCustomer.html',{'dcust':dc})
												except	Customers.DoesNotExist:																																	
																perror="Given Customer Mobile number Doesn't exist"
									elif  'dcust'	in	str(request.POST):
												dc=get_object_or_404(Customers,	pk=str(request.POST['dcust']).strip())
												dc.delete()
												psuccess="Successfully Deleted the Customer"
											
						return render(request,'products/DeleteCustomer.html',{'perror':perror,'psuccess':psuccess})


@login_required
def add_sales(request):
		return render(request,'products/AddSalesItem.html',{})


@login_required
def edit_sales(request):
		return render(request,'products/EditSalesItem.html',{})


@login_required
def delete_sales(request):
		return render(request,'products/DeleteSalesItem.html',{})

@login_required
def cart_checkout(request):
		perror=""
		psuccess=""
		size=20
		order=""
		chars=string.ascii_lowercase + string.digits
		rdm='ISKCONGNT'+''.join(random.choice(chars) for _ in range(size))
		cust=""
		trid=""
		cash=0
		online=0
		gst=0
		iprice=0
		order_price=0
		paid=0
		gift=False
		damage=False
		if Cart.objects.exists():
			for j in Cart.objects.all():
				gst=gst+(j.tax*j.order_quantity)
				iprice=iprice+(j.item_price*j.order_quantity)
			gst=round(gst,2)
			iprice=round(iprice,2)
			order_price=round(Cart.objects.aggregate(Sum('order_price'))['order_price__sum'],2)
		#import pdb; pdb.set_trace()
		#gift=str(request.POST['gift']).strip()
		if request.POST and str(request.POST['mobile']).strip():
			if str(request.POST['gift']).strip()=="Y" and str(request.POST['damage']).strip()=="N":
				gift=True
				paid=1
			elif str(request.POST['damage']).strip()=="Y" and str(request.POST['gift']).strip()=="N":
				damage=True
				paid=1
			elif str(request.POST['damage']).strip()=="Y" and str(request.POST['gift']).strip()=="Y":
				perror=" Order can be either Damage order or Gift order but not both"
				return render(request,'products/CartCheckOut.html',{'cust':cust,'cart':Cart.objects.all().order_by('bu'),'tax':gst,'iprice':iprice,'total':order_price,'psuccess':psuccess,'perror':perror})
			else:
					if str(request.POST['cash']).strip():
						cash=round(float(str(request.POST['cash']).strip()),2)
									
					if str(request.POST['online']).strip():
						online=round(float(str(request.POST['online']).strip()),2)

					if online > 0:
						trid=str(request.POST['trid']).strip()
						if trid == "":
							perror=" Please enter Transaction ID for online payment"
							return render(request,'products/CartCheckOut.html',{'cust':cust,'cart':Cart.objects.all().order_by('bu'),'tax':gst,'iprice':iprice,'total':order_price,'psuccess':psuccess,'perror':perror})
					total_paid	= cash + online
					if total_paid==order_price:
						paid=1
					else:
						perror="Order amount should be paid either by cash or through online"
						return render(request,'products/CartCheckOut.html',{'cust':cust,'cart':Cart.objects.all().order_by('bu'),'tax':gst,'iprice':iprice,'total':order_price,'psuccess':psuccess,'perror':perror})

			if paid==1:
				try:
					cust= Customers.objects.get(pk=str(request.POST['mobile']).strip())
				except	Customers.DoesNotExist:
							mobile=request.POST.get('mobile')
							name=request.POST.get('name')
							mail=request.POST.get('mail')
							addr=request.POST.get('address')
							cust= Customers(customer_mobile=mobile,customer_name=name,customer_mailid=mail,customer_address=addr)
							cust.save()
				order=Orders(order_id=rdm,order_status='Success',order_date=timezone.now(),order_total=order_price,item_price=iprice,tax=gst,cash=cash,online=online,gift=gift,transaction_id=trid,damage=damage,customer_mobile=Customers.objects.get(pk=cust.customer_mobile))
				order.save()
				for item in Cart.objects.all():
					ist=ItemSales(item_code=item.item_code,item_name=item.name,unit=item.unit,bu=item.bu,group=item.item_group,order_quantity=item.order_quantity,item_price=item.item_price,igst=item.igst,cgst=item.cgst,sgst=item.sgst,tax=item.tax,item_net_price=item.item_net_price,order_price=item.order_price,damage=item.damage,customer_mobile=Customers.objects.get(pk=cust.customer_mobile),order_id=Orders.objects.get(pk=order.order_id))
					ist.save()
					item.delete()
				qs=ItemSales.objects.all()
				qs=qs.filter(order_id=order.order_id).order_by('bu')
				return render(request,'products/ConfirmOrder.html',{'cust':cust,'order':order,'qs':qs,'total':round(qs.aggregate(Sum('order_price'))['order_price__sum'],2),'psuccess':psuccess,'perror':perror})
			
				
		return render(request,'products/CartCheckOut.html',{'cust':cust,'cart':Cart.objects.all().order_by('bu'),'tax':gst,'iprice':iprice,'total':order_price,'psuccess':psuccess,'perror':perror})

#@login_required
#def generatePDF11(request):
	#buf=io.BytesIO()
	#c=canvas.Canvas(buf,pagesize=A5, bottomup=0)
	#c.setFont(psfontname, size, leading = None)
	#c.setFont("Helvetica",14)
	#c.showOutline()

	#fname = os.path.join(BASE_DIR, 'pdf_test_%s.pdf' %
	#                    ( datetime.now().strftime('%d-%b-%Y %H-%M')))
	#fname = 'pdf_test_%s.pdf'%( datetime.datetime.now().strftime('%d-%b-%Y %H-%M'))
	#data=[['x','y','z'],['1','2','3'],['4','5','6'],['7','8','9']]

	
	#width = 400
	#height = 100
	#x = 100
	#y = 800
	#f = Table(data)
	#textob=c.beginText()
	#textob.setTextOrigin(inch,inch)
	#textob.setFont("Helvetica",14)
	#lines=['line1', 'line2', 'line3']
	#for line in lines:
	#			textob.textLine(line)

	#f.setStyle(TableStyle([('TEXTCOLOR', (0, 0), (1, -1), colors.red)]))

	#f.wrapOn(c, width, height)
	#f.drawOn(c, x, y)

	#doc = SimpleDocTemplate(fname)
	#elements = [] 
	#elements.append(f) 
	#doc.build(elements)

	#c.drawText(textob)
	#c.showPage()
	#c.save()
	#buf.seek(0)	

	#mydoc=SimpleDocTemplate(fname,pagesize=A5)
	#data=[['x','y','z'],['1','2','3'],['4','5','6'],['7','8','9']]
	#t=Table(data)
	#elements=[]
	#elements.append(t)
	#mydoc.build(elements)
	
	#c.(mydoc)
	#c.showPage()
	#c.save()
	#return FileResponse(buf, as_attachment=True, filename=fname)

	#response = HttpResponse(content_type='application/pdf') 
	#buffer=io.BytesIO()
	#x=canvas.Canvas(response)
	
	#data=[['x','y','z'],['1','2','3'],['4','5','6'],['7','8','9']]
	
	#filename = os.path.join(BASE_DIR, 'pdf_test_%s.pdf' %
	#                    ( datetime.now().strftime('%d-%b-%Y %H-%M')))

	#pdf= SimpleDocTemplate(filename,pagesize=letter)
	#table=Table(data)
	##elems=[]
	#elems.append(table)
	#pdf.build(elems)

	#x.drawString(100,100,pdf)
	#x.showPage()
	#x.save()
	#buffer.seek(0)
	#return FileResponse(buffer,as_attachment=True, filename='attempt1.pdf')
	#return response



@login_required
def gen_report(request):
		qs=""
		#qs=Orders.objects.all()
		psuccess=""
		perror=""
		total_tax=0
		total_order=0
		o_dict={}
		gift=False
		gift_val=""
		damage=False
		damage_vl=""
		qs=""
		#import pdb; pdb.set_trace()
		if request.POST:
			#import pdb; pdb.set_trace()
			o_select=str(request.POST.get('o_select')).strip()
			d_from=str(request.POST.get('d_from')).strip()
			d_to=str(request.POST.get('d_to')).strip()

			qs=Orders.objects.all()
			
			if (str(request.POST.get('gift')).strip() =="Y"):
				gift_val=True
			elif (str(request.POST.get('gift')).strip() =="N"):
				gift_val=False
			if (str(request.POST.get('damage')).strip() =="Y"):
				damage_val=True
			elif (str(request.POST.get('damage')).strip() =="N"):
				damage_val=False


			if (str(request.POST.get('gift')).strip() =="Y") | (str(request.POST.get('gift')).strip() =="N"):
				qs=qs.filter(gift=gift_val)
			if (str(request.POST.get('damage')).strip() =="Y") | (str(request.POST.get('damage')).strip() =="N"):
				qs=qs.filter(damage=damage_val)
				
			today=datetime.date.today()
			# qs=""
			if o_select== 'Today':
				# if gift:
				# 	qs=Orders.objects.filter(Q(order_date__date=datetime.datetime.strptime(str(today),'%Y-%m-%d').date()) & Q(gift=gift_val))
				# else:
				#	qs=Orders.objects.filter(order_date__date=datetime.datetime.strptime(str(today),'%Y-%m-%d').date())

				qs=qs.filter(order_date__date=datetime.datetime.strptime(str(today),'%Y-%m-%d').date())
			elif o_select=='Specific Date':
				# if gift:
					if d_from :
						#qs=Orders.objects.filter(Q(order_date__date=datetime.datetime.strptime(d_from,'%Y-%m-%d').date())& Q(gift=gift_val))
						qs=qs.filter(Q(order_date__date=datetime.datetime.strptime(d_from,'%Y-%m-%d').date()))
					else:
						perror="Pls select the from date"

				# else:
				# 	if d_from :
				# 		qs=Orders.objects.filter(order_date__date=datetime.datetime.strptime(d_from,'%Y-%m-%d').date())
				# 	else:
				# 		perror="Pls select the from date"

			elif o_select=='Last 7 Days':
				# if gift:
				# 	qs=Orders.objects.filter(Q(order_date__gte=datetime.datetime.now()-timedelta(days=7)) & Q(gift=gift_val))
				# else:
					# qs=Orders.objects.filter(order_date__gte=datetime.datetime.now()-timedelta(days=7))
					qs=qs.filter(order_date__gte=datetime.datetime.now()-timedelta(days=7))

			elif o_select=='Last 15 Days':
				# if gift:
				# 	qs=Orders.objects.filter(Q(order_date__gte=datetime.datetime.now()-timedelta(days=15)) & Q(gift=gift_val))
				# else:
				# 	qs=Orders.objects.filter(order_date__gte=datetime.datetime.now()-timedelta(days=15))
					qs=qs.filter(order_date__gte=datetime.datetime.now()-timedelta(days=15))

			elif o_select=='Date Range':
				
				if d_from and d_to:
					#import pdb; pdb.set_trace()
					#if datetime.datetime.strptime(d_from,'%Y-%m-%d').date() > datetime.datetime.strptime(d_to,'%Y-%m-%d').date(): 
					if d_from > d_to:
						perror=" From date should be less than To date"
					else:
						# if gift:
						# 	qs=Orders.objects.filter(Q(order_date__range=[datetime.datetime.strptime(d_from,'%Y-%m-%d').date(), datetime.datetime.strptime(d_to,'%Y-%m-%d').date()])&Q(gift=gift_val))
						# else:
						# 	qs=Orders.objects.filter(order_date__range=[datetime.datetime.strptime(d_from,'%Y-%m-%d').date(), datetime.datetime.strptime(d_to,'%Y-%m-%d').date()])
						qs=qs.filter(order_date__range=[datetime.datetime.strptime(d_from,'%Y-%m-%d').date(), datetime.datetime.strptime(d_to,'%Y-%m-%d').date()])
				else:
					perror="Pls select Both from and to dates"

			o_dict={}
			#import pdb; pdb.set_trace()
			request.session['o_data']=""
			j=[]
			for i in qs:
				o_dict[i]=ItemSales.objects.filter(order_id=i.order_id)
				j.append(i.order_id)
			request.session['o_data']=j
			#import pdb; pdb.set_trace()
			if qs: 
				total_tax=round(qs.aggregate(Sum('tax'))['tax__sum'],2)
				total_order=round(qs.aggregate(Sum('order_total'))['order_total__sum'],2)
			#request.session['data']=o_dict
		return render(request,'products/Reports.html',{'o_dict':o_dict,'total_tax':total_tax,'total_order':total_order,'psuccess':psuccess,'perror':perror})



@login_required
def generatePDF(request):
		response=""
		if request.POST:
				#import pdb; pdb.set_trace()
				o_id=str(request.POST.get('o_id')).strip()
				order=Orders.objects.get(pk=o_id)
				qs=ItemSales.objects.filter(order_id=o_id).order_by('bu')
				fname = 'order_%s.pdf'%( datetime.datetime.now().strftime('%d-%b-%Y %H-%M'))
				template_path = 'products/PrintOrder.html'
				context = {'cust':order.customer_mobile,'order':order,'qs':qs,'total':qs.aggregate(Sum('order_price'))['order_price__sum']}
				# Create a Django response object, and specify content_type as pdf
				response = HttpResponse(content_type='application/pdf')
				response['Content-Disposition'] = 'attachment; filename='+fname
				# find the template and render it.
				template = get_template(template_path)
				html = template.render(context)

				# create a pdf
				pisa_status = pisa.CreatePDF(
				html, dest=response,link_callback=link_callback )
				# if error then show some funny view
				if pisa_status.err:
					return HttpResponse('We had some errors <pre>' + html + '</pre>')
		return response


def link_callback(uri, rel):
			"""
			Convert HTML URIs to absolute system paths so xhtml2pdf can access those
			resources
			"""

			result = finders.find(uri)
			if result:
					if not isinstance(result, (list, tuple)):
							result = [result]
					result = list(os.path.realpath(path) for path in result)
					path=result[0]
			else:
					sUrl = settings.STATIC_URL        # Typically /static/
					sRoot = settings.STATIC_ROOT      # Typically /home/userX/project_static/
					mUrl = settings.MEDIA_URL         # Typically /media/
					mRoot = settings.MEDIA_ROOT       # Typically /home/userX/project_static/media/

					if uri.startswith(mUrl):
							path = os.path.join(mRoot, uri.replace(mUrl, ""))
					elif uri.startswith(sUrl):
							path = os.path.join(sRoot, uri.replace(sUrl, ""))
					else:
							return uri

			# make sure that file exists
			if not os.path.isfile(path):
					raise Exception(
							'media URI must start with %s or %s' % (sUrl, mUrl)
					)
			return path

@login_required
def generatePDF1(request):
		response=""
		o_dict={}
		o_dict_list=[]
		total_tax=0
		total_order=0
		if request.GET:
				#import pdb; pdb.set_trace()
				j=1
				
				o_dict_list=request.session['o_data']

				for i in o_dict_list:
					order=Orders.objects.get(pk=i)
					o_dict[order]=ItemSales.objects.filter(order_id=i)
					total_tax=total_tax+order.tax
					total_order=total_order+order.order_total
				total_tax=round(total_tax,2)	
				total_order=round(total_order,2)
				
				#request.session['data']=o_dict
				#return render(request,'products/Reports.html',{'o_dict':o_dict,'total_tax':total_tax,'total_order':total_order,'psuccess':psuccess,'perror':perror})

				
				fname = 'Report_%s.pdf'%( datetime.datetime.now().strftime('%d-%b-%Y %H-%M'))
				template_path = 'products/PrintOrders.html'
				context = {'o_dict':o_dict,'total_tax':total_tax,'total_order':total_order}
				# Create a Django response object, and specify content_type as pdf
				response = HttpResponse(content_type='application/pdf')
				response['Content-Disposition'] = 'attachment; filename='+fname
				# find the template and render it.
				template = get_template(template_path)
				html = template.render(context)

				# create a pdf
				pisa_status = pisa.CreatePDF(
				html, dest=response,link_callback=link_callback )
				# if error then show some funny view
				if pisa_status.err:
					return HttpResponse('We had some errors <pre>' + html + '</pre>')
		return response


def gen_summary(request):
		qs=""
		#qs=Orders.objects.all()
		psuccess=""
		perror=""
		total_tax=0
		total_order=0
		total_cash=0
		total_online=0
		total_gifts=0
		bu_dict={}
		#import pdb; pdb.set_trace()
		if request.POST:
			#import pdb; pdb.set_trace()
			o_select=str(request.POST.get('o_select')).strip()
			d_from=str(request.POST.get('d_from')).strip()
			d_to=str(request.POST.get('d_to')).strip()
			today=datetime.date.today()
			qs=""
			if o_select== 'Today':
				qs=Orders.objects.filter(order_date__date=datetime.datetime.strptime(str(today),'%Y-%m-%d').date())

			elif o_select=='Specific Date':
				if d_from :
					qs=Orders.objects.filter(order_date__date=datetime.datetime.strptime(d_from,'%Y-%m-%d').date())
				else:
					perror="Pls select the from date"
			elif o_select=='Last 7 Days':
				qs=Orders.objects.filter(order_date__gte=datetime.datetime.now()-timedelta(days=7))

			elif o_select=='Last 15 Days':
				qs=Orders.objects.filter(order_date__gte=datetime.datetime.now()-timedelta(days=15))

			elif o_select=='Date Range':
				
				if d_from and d_to:
					#import pdb; pdb.set_trace()
					#if datetime.datetime.strptime(d_from,'%Y-%m-%d').date() > datetime.datetime.strptime(d_to,'%Y-%m-%d').date(): 
					if d_from > d_to:
						perror=" From date should be less than To date"
					else:
						qs=Orders.objects.filter(order_date__range=[datetime.datetime.strptime(d_from,'%Y-%m-%d').date(), datetime.datetime.strptime(d_to,'%Y-%m-%d').date()])
				else:
					perror="Pls select Both from and to dates"

			request.session['o_summary']=""
			order_list=[]
			total_gifts=0
			if qs :
				for m in qs:
					order_list.append(m.order_id)
					if m.gift:
						total_gifts=total_gifts+m.order_total
					
				#order_list=qs.values_list('order_id',flat=True)
				request.session['o_summary']=order_list
				total_items=ItemSales.objects.filter(order_id__in=qs)
				bu_list=total_items.values_list('bu',flat=True).distinct()
				
				bu_dict={}
				for i in bu_list:
					bu_tax=0
					bu_price=0
					#import pdb; pdb.set_trace()
					bu_tax=round(total_items.filter(bu=i).aggregate(Sum('tax'))['tax__sum'],2)
					bu_price=round(total_items.filter(bu=i).aggregate(Sum('order_price'))['order_price__sum'],2)
					grp_list=[]
					bu_items=total_items.filter(bu=i)
					grp_list=bu_items.values_list('group',flat=True).distinct()
					grp_dict={}
					for j in grp_list:
						grp_items=bu_items.filter(group=j)
						grp_tax=0
						grp_price=0
						grp_tax=round(grp_items.aggregate(Sum('tax'))['tax__sum'],2)
						grp_price=round(grp_items.aggregate(Sum('order_price'))['order_price__sum'],2)
						grp_dict[j]=[grp_tax,grp_price]
					bu_dict[i]={'tax':bu_tax,
								'price':bu_price,
								'groups':grp_dict}
				total_tax=round(qs.aggregate(Sum('tax'))['tax__sum'],2)
				total_order=round(qs.aggregate(Sum('order_total'))['order_total__sum'],2)
				total_cash=round(qs.aggregate(Sum('cash'))['cash__sum'],2)
				total_online=round(qs.aggregate(Sum('online'))['online__sum'],2)
		#import pdb; pdb.set_trace()
		return render(request,'products/SummaryReport.html',{'bu_dict':bu_dict,'total_tax':total_tax,'total_order':total_order,'total_cash':total_cash,'total_online':total_online,'total_gifts':total_gifts,'psuccess':psuccess,'perror':perror})

@login_required
def gen_summary_repo(request):
		response=""
		o_dict={}
		total_tax=0
		total_order=0
		total_cash=0
		total_online=0
		total_gifts=0
		o_list=[]

		if request.GET:
				o_list=request.session['o_summary']

				for i in o_list:
					order=Orders.objects.get(pk=i)
					o_dict[order]=ItemSales.objects.filter(order_id=i)
					total_tax=total_tax+order.tax
					total_order=total_order+order.order_total
					total_cash=total_cash+order.cash
					total_online=total_online+order.online
					#import pdb; pdb.set_trace()
					if order.gift:
						#import pdb; pdb.set_trace()
						total_gifts=total_gifts+order.order_total

				total_tax=round(total_tax,2)	
				total_order=round(total_order,2)
				total_cash=round(total_cash,2)
				total_online=round(total_online,2)
				total_gifts=round(total_gifts,2)
				total_items=ItemSales.objects.filter(order_id__in=o_list)
				bu_list=total_items.values_list('bu',flat=True).distinct()
				#import pdb; pdb.set_trace()
				bu_dict={}
				for i in bu_list:
					bu_tax=0
					bu_price=0
					#import pdb; pdb.set_trace()
					bu_tax=round(total_items.filter(bu=i).aggregate(Sum('tax'))['tax__sum'],2)
					bu_price=round(total_items.filter(bu=i).aggregate(Sum('order_price'))['order_price__sum'],2)
					grp_list=[]
					bu_items=total_items.filter(bu=i)
					grp_list=bu_items.values_list('group',flat=True).distinct()
					grp_dict={}
					for j in grp_list:
						grp_items=bu_items.filter(group=j)
						grp_tax=0
						grp_price=0
						grp_tax=round(grp_items.aggregate(Sum('tax'))['tax__sum'],2)
						grp_price=round(grp_items.aggregate(Sum('order_price'))['order_price__sum'],2)
						grp_dict[j]=[grp_tax,grp_price]
					bu_dict[i]={'tax':bu_tax,
								'price':bu_price,
								'groups':grp_dict}


				#request.session['data']=o_dict
				#return render(request,'products/Reports.html',{'o_dict':o_dict,'total_tax':total_tax,'total_order':total_order,'psuccess':psuccess,'perror':perror})

				
				fname = 'Summary_%s.pdf'%( datetime.datetime.now().strftime('%d-%b-%Y %H-%M'))
				template_path = 'products/PrintSummary.html'
				context = {'bu_dict':bu_dict,'total_tax':total_tax,'total_order':total_order,'total_cash':total_cash,'total_online':total_online,'total_gifts':total_gifts}
				# Create a Django response object, and specify content_type as pdf
				response = HttpResponse(content_type='application/pdf')
				response['Content-Disposition'] = 'attachment; filename='+fname
				# find the template and render it.


				template = get_template(template_path)
				html = template.render(context)

				# create a pdf
				pisa_status = pisa.CreatePDF(
				html, dest=response,link_callback=link_callback )
				# if error then show some funny view
				if pisa_status.err:
					return HttpResponse('We had some errors <pre>' + html + '</pre>')
		#import pdb; pdb.set_trace()
		#return render(request,'products/SummaryReport.html',{'bu_dict':bu_dict,'total_tax':total_tax,'total_order':total_order,'total_cash':total_cash,'total_online':total_online,'psuccess':psuccess,'perror':perror})

		return response
@login_required
def list_alerts(request):
		#import pdb; pdb.set_trace()
		za=InventoryItems.objects.filter(available_quantity=0).order_by('bu')
		li=InventoryItems.objects.filter(Q(available_quantity__gte=1) & Q(available_quantity__lte=5)).order_by('bu')
		#import pdb; pdb.set_trace()
		ec=InventoryItems.objects.filter(Q(exp_date__gte=(datetime.datetime.now()-timedelta(days=30))) & Q(exp_date__lte=datetime.datetime.now()))
		exp=InventoryItems.objects.filter(exp_date__lte=datetime.datetime.now())
		damage=InventoryItems.objects.filter(damage=True).order_by('bu')
		za_dict={}
		li_dict={}
		ec_dict={}
		exp_dict={}
		damage_dict={}
		if za:
			za_bu=[]
			za_bu=za.values_list('bu',flat=True).distinct()
			za_dict={}
			if za_bu:
				for i in za_bu:
					qs=""
					qs=za.filter(bu=BusinessUnit.objects.get(pk=i))
					za_dict[i]=qs
		if li:
			li_bu=[]
			li_bu=li.values_list('bu',flat=True).distinct()
			li_dict={}
			if li_bu:
				for j in li_bu:
					qs=""
					qs=li.filter(bu=BusinessUnit.objects.get(pk=j))
					li_dict[j]=qs
		if ec:
			ec_bu=[]
			ec_bu=ec.values_list('bu',flat=True).distinct()
			ec_dict={}
			if ec_bu:
				for k in ec_bu:
					qs=""
					qs=ec.filter(bu=BusinessUnit.objects.get(pk=k))
					ec_dict[k]=qs
		
		if exp:
			exp_bu=[]
			exp_bu=exp.values_list('bu',flat=True).distinct()
			exp_dict={}
			if exp_bu:
				for l in exp_bu:
					qs=""
					qs=exp.filter(bu=BusinessUnit.objects.get(pk=l))
					exp_dict[l]=qs

		if damage:
			damage_bu=[]
			damage_bu=damage.values_list('bu',flat=True).distinct()
			damage_dict={}
			if damage_bu:
				for i in damage_bu:
					qs=""
					qs=damage.filter(bu=BusinessUnit.objects.get(pk=i))
					damage_dict[i]=qs
		return render(request,'products/ListAlerts.html',{'za_dict':za_dict,'li_dict':li_dict,'ec_dict':ec_dict,'exp_dict':exp_dict, 'damage_dict':damage_dict})




@login_required
def get_inventory(request):

		za=InventoryItems.objects.order_by('bu','group')
		#import pdb; pdb.set_trace()
		za_dict={}
		if za:
			za_bu=[]
			za_bu=za.values_list('bu',flat=True).distinct()
			za_dict={}
			if za_bu:
				for i in za_bu:
					qs=""
					qs=za.filter(bu=BusinessUnit.objects.get(pk=i))
					za_dict[i]=qs
	
		return render(request,'products/ListInventoryCurrentStatus.html',{'za_dict':za_dict})


@login_required
def gen_alert_repo(request):
		response=""
		#import pdb; pdb.set_trace()
		za_dict={}
		li_dict={}
		ec_dict={}
		exp_dict={}
		damage_dict={}
		if request.GET:

				za=InventoryItems.objects.filter(available_quantity=0).order_by('bu')
				li=InventoryItems.objects.filter(Q(available_quantity__gte=1) & Q(available_quantity__lte=5)).order_by('bu')
				ec=InventoryItems.objects.filter(Q(exp_date__gte=(datetime.datetime.now()-timedelta(days=30))) & Q(exp_date__lte=datetime.datetime.now()))
				exp=InventoryItems.objects.filter(exp_date__lte=datetime.datetime.now())
				damage=InventoryItems.objects.filter(damage=True).order_by('bu')

				if za:
					za_bu=[]
					za_bu=za.values_list('bu',flat=True).distinct()
					za_dict={}
					if za_bu:
						for i in za_bu:
							qs=""
							qs=za.filter(bu=BusinessUnit.objects.get(pk=i))
							za_dict[i]=qs
				if li:
					li_bu=[]
					li_bu=li.values_list('bu',flat=True).distinct()
					li_dict={}
					if li_bu:
						for j in li_bu:
							qs=""
							qs=li.filter(bu=BusinessUnit.objects.get(pk=j))
							li_dict[j]=qs
				if ec:
					ec_bu=[]
					ec_bu=ec.values_list('bu',flat=True).distinct()
					ec_dict={}
					if ec_bu:
						for k in ec_bu:
							qs=""
							qs=ec.filter(bu=BusinessUnit.objects.get(pk=k))
							ec_dict[k]=qs
				
				if exp:
					exp_bu=[]
					exp_bu=exp.values_list('bu',flat=True).distinct()
					exp_dict={}
					if exp_bu:
						for l in exp_bu:
							qs=""
							qs=exp.filter(bu=BusinessUnit.objects.get(pk=l))
							exp_dict[l]=qs

				if damage:
					damage_bu=[]
					damage_bu=damage.values_list('bu',flat=True).distinct()
					damage_dict={}
					if damage_bu:
						for i in damage_bu:
							qs=""
							qs=damage.filter(bu=BusinessUnit.objects.get(pk=i))
							damage_dict[i]=qs
								
				fname = 'Alert_%s.pdf'%( datetime.datetime.now().strftime('%d-%b-%Y %H-%M'))

				template_path = 'products/PrintAlerts.html'
				context = {'za_dict':za_dict,'li_dict':li_dict,'ec_dict':ec_dict,'exp_dict':exp_dict, 'damage_dict':damage_dict}
				# Create a Django response object, and specify content_type as pdf
				response = HttpResponse(content_type='application/pdf')
				response['Content-Disposition'] = 'attachment; filename='+fname
				# find the template and render it.
				template = get_template(template_path)
				html = template.render(context)

				# create a pdf
				pisa_status = pisa.CreatePDF(
				html, dest=response,link_callback=link_callback )
				# if error then show some funny view
				if pisa_status.err:
					return HttpResponse('We had some errors <pre>' + html + '</pre>')
		return response



@login_required
def gen_inventory_repo(request):
		response=""
		#import pdb; pdb.set_trace()
		za_dict={}

		if request.GET:

				za=InventoryItems.objects.order_by('bu','group')

				if za:
					za_bu=[]
					za_bu=za.values_list('bu',flat=True).distinct()
					za_dict={}
					if za_bu:
						for i in za_bu:
							qs=""
							qs=za.filter(bu=BusinessUnit.objects.get(pk=i))
							za_dict[i]=qs
				

				fname = 'Alert_%s.pdf'%( datetime.datetime.now().strftime('%d-%b-%Y %H-%M'))

				template_path = 'products/PrintCurrentInventory.html'
				context = {'za_dict':za_dict}
				# Create a Django response object, and specify content_type as pdf
				response = HttpResponse(content_type='application/pdf')
				response['Content-Disposition'] = 'attachment; filename='+fname
				# find the template and render it.
				template = get_template(template_path)
				html = template.render(context)

				# create a pdf
				pisa_status = pisa.CreatePDF(
				html, dest=response,link_callback=link_callback )
				# if error then show some funny view
				if pisa_status.err:
					return HttpResponse('We had some errors <pre>' + html + '</pre>')
		return response


@login_required
def list_orders(request):
		qs=""
		order_id=request.GET.get('order_id')
		order_status=request.GET.get('order_status')
		order_date=request.GET.get('order_date')
		order_total=request.GET.get('order_total')
		order_customer=request.GET.get('order_customer')
				
		if order_id or order_status or order_date or order_total or order_customer:
				qs=Orders.objects.all()

				if order_id !="" and order_id is not None:
						qs=qs.filter(order_id__icontains=order_id)
				if order_status !="" and order_status is not None:
						qs=qs.filter(order_status__icontains=order_status)
				if order_date !="" and order_date is not None:
						qs=qs.filter(order_date__date=datetime.datetime.strptime(order_date,'%Y-%m-%d').date())
				if order_total !="" and order_total is not None:
						qs=qs.filter(order_total__icontains=order_total)
				if order_customer !="" and order_customer is not None:
						try: 
							qs=qs.filter(customer_mobile=Customers.objects.get(pk=str(order_customer).strip())	)
						except Customers.DoesNotExist:
							qs=qs
				if len(qs)==1:
					#import pdb; pdb.set_trace()
					cust= qs[0].customer_mobile
					qs1=ItemSales.objects.filter(order_id=qs[0].order_id).order_by('bu')
					return render(request,'products/ConfirmOrder.html',{'cust':cust,'order':qs[0],'qs':qs1,'total':round(qs1.aggregate(Sum('order_price'))['order_price__sum'],2)})

					

		return render(request,'products/ListOrders.html',{'orders':qs})

def trial(request):
    call_command('dbbackup')
    return HttpResponse('db created')

@login_required
def return_items(request):
	qs=[]
	qs1=[]

	order_id=""
	perror=""
	order_id=request.GET.get('order_id')

	if order_id !="" and order_id is not None:
			qs=Orders.objects.filter(order_id__icontains=order_id)
			qs1=ItemSales.objects.filter(order_id=Orders.objects.get(pk=str(order_id).strip()))

	if request.POST:
			qs1=[]
			#import pdb; pdb.set_trace()
			o_id=str(request.POST.get('order_id')).strip()

			qty_edited=[]
			if o_id != "" and o_id is not None:
				updated_order=Orders.objects.filter(order_id=o_id)
				qs1=ItemSales.objects.filter(order_id=Orders.objects.get(pk=str(o_id).strip()))
				qty_edited=request.POST.getlist('quntity')
				if len(qty_edited) == len(qs1):
					i=0
					upd_o_total=updated_order.order_total
					upd_o_tax=updated_order.tax
					upd_o_item_price=updated_order.item_price

					for item in qs1:
						inv_item=""
						updated_order=""
						old_item_order_price=""
						if qty_edited[i] != "":
							if int(qty_edited[i]) < item.order_quantity  :
								old_item_order_price=item.order_price
								item.order_quantity = qty_edited[i]
								item.order_price=(item.order_quantity*item.item_net_price)
								#item.save()
								difference_order_price+=(old_item_order_price-item.order_price)



								inv_item=InventoryItems.objects.filter(item_code=item.item_code)
								inv_item.available_quantity += (item.order_quantity - int(qty_edited[i]))
								#inv_item.save()

								#updated_order.

							elif int(qty_edited[i]) > item.order_quantity:
									perror="Edited Order quantity can't be more than what was ordered before"
									break
						i=i+1

				else:
					perror="There is some issue with the Return of goods, pls check with Admin"


	return render(request,'products/ReturnOfItems.html',{'orders':qs,'order_items':qs1, 'perror':perror})


@login_required
def list_order_items(request):
		qs=[]
		order_id=request.GET.get('order_id')
		
		if order_id :
			qs=ItemSales.objects.all()
			if order_id !="" and order_id is not None:
						qs=qs.filter(order_id=Orders.objects.get(pk=str(order_id).strip()))

		return render(request,'products/ListOrderedItems.html',{'order_items':qs})

@login_required
def import_items(request):
				success=""
				error=""
				exclusions=[]
				inclusions=[]
				buexclusions=[]
				igexclusions=[]
				updated=[]

				Templates={
				'ITEMS':['ITEM_CODE','BAR_CODE','ITEM_NAME','UPD','DAMAGE','UNIT','QTY','PRICE','EXP_DATE','IGST','CGST','SGST','BU','ITEM_GROUP'],
				
				}
				if	request.POST:
																				rfile1=request.FILES['excel_file']
																				if	not	rfile1.name.endswith('xlsx'):
																							error="Pls	upload	the	xlsx	file	only	"

																				else:
																							
																							today=timezone.now()

																							wb = openpyxl.load_workbook(rfile1)
																											
																							if	'ITEMS'	not	in	wb.sheetnames:
																											error=	"There	is	no	Sheet	with	the	name	ITEMS"
																							else:	
																										ws	=	wb['ITEMS']
																										
																										collist=Templates['ITEMS']
																										first_row = []
																										flag=0
																										i=0
																										while	i	<	len(collist):
																														if	collist[i]!=str(ws.cell(row=1,column=(i+1)).value).strip():
																																	flag=1
																																	break
																														i+=1
																										if	flag==0:
																														
																														
																														bu_list=[]
																														bu_list=BusinessUnit.objects.values_list('name',flat=True)
																														

																														mxrows = (ws.max_row)-1
																														count=0
																														for	row	in ws.iter_rows(min_row=2):
																																		if	row[0].value	is	not	None:
																																						grp_list=[]
																																						bu_val=""
																																						ig_val=""
																																						
																																						row_data=[]
																																						
																																						for	cell	in	row:
																																										x=str(cell.value).strip()
																																										if	cell.value	is	None:
																																												x=""																																																		
																																										row_data.append(x)
																																						bu_val=row_data[12].strip()
																																						ig_val=row_data[13].strip()
																																						#import pdb; pdb.set_trace();
																																						if bu_val != "":
																																							grp_list=InventoryGroups.objects.filter(bu=BusinessUnit(name=bu_val)).values_list('name',flat=True)
																																							if grp_list != "":
																																								grp_list=grp_list.values_list('name',flat=True)

																																						if	str(row_data[0])	!=	"":
																																										#import pdb; pdb.set_trace()
																																										exp_date="1111-11-11"
																																										damage=False
																																										if row_data[8].strip()=="":
																																											exp_date="1111-11-11"
																																										else:
																																											exp_date=row_data[8].strip()
																																										
																																										if row_data[4].strip()=='Y':
																																											damage=True

																																										try:
																																																obj = InventoryItems.objects.get(pk=row_data[0])
																																																if row_data[3]=="Y":
																																																	obj.name=row_data[2]
																																																	obj.damage=damage
																																																	obj.unit=row_data[5]
																																																	obj.available_quantity=row_data[6]
																																																	obj.price=row_data[7]
																																																	obj.exp_date=exp_date
																																																	obj.igst=row_data[9]
																																																	obj.cgst=row_data[10]
																																																	obj.sgst=row_data[11]
																																																	obj.bu=BusinessUnit(name=row_data[12].strip(),desc=row_data[12].strip(),date=today)
																																																	obj.group=InventoryGroups(name=row_data[13].strip(),desc=row_data[13].strip(),date=today,bu=BusinessUnit(name=row_data[12].strip(),desc=row_data[12].strip(),date=today))
																																																	obj.save()
																																																	updated.append({'Item':row_data[0],'InventoryGroup':row_data[13].strip(),'BusinessUnit':row_data[12].strip()})


																																										except	InventoryItems.DoesNotExist:
																																																	
																																																	#import pdb; pdb.set_trace()
																																																	if	bu_val	==	""	or	bu_val	not	in	bu_list:
																																																				buexclusions.append({'Item':row_data[0],'BusinessUnit':bu_val})
																																																	elif ig_val	== "" or ig_val not	in	grp_list:
																																																					igexclusions.append({'Item':row_data[0],'InventoryGroup':ig_val, 'BusinessUnit':bu_val })
																																																	else:
																																																		#import pdb; pdb.set_trace()
																																																		#today=datetime.date.today()
																																																		obj=InventoryItems(
																																																		item_code=row_data[0].strip(),
																																																		name=(row_data[2].strip()),
																																																		damage=damage,
																																																		unit=row_data[5].strip(),
																																																		available_quantity=row_data[6].strip(),
																																																		price=row_data[7].strip(),
																																																		exp_date=exp_date,
																																																		igst=row_data[9].strip(),
																																																		cgst=row_data[10].strip(),
																																																		sgst=row_data[11].strip(),
																																																		bu=BusinessUnit(name=row_data[12].strip(),desc=row_data[12].strip(),date=today),
																																																		group=InventoryGroups(name=row_data[13].strip(),desc=row_data[13].strip(),date=today,bu=BusinessUnit(name=row_data[12].strip(),desc=row_data[12].strip(),date=today)),
																																																		date=today

																																																		)
																																																		obj.save()
																																																		inclusions.append({'Item':row_data[0],'InventoryGroup':row_data[13].strip(),'BusinessUnit':row_data[12].strip()})
																																		#count+=1
																																		#job = updateTask.delay(int((count / mxrows)*100))
																																		##task_status(job.id)
																																		#return render(request,'hwt/ImportData.html',{'task_id':job.id})													

																										else:
																													error=	"Uploaded	file	is	not	having	right	columns"
				
				return render(request,'products/ImportItems.html',{'success':success,'error':error,'inclusions':inclusions,'buexclusions':buexclusions,'igexclusions':igexclusions,'updated':updated})

		

